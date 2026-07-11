#!/usr/bin/env python3
"""
fleet_stats.py v2 — auto-built fleet performance workbook.

Data sources (all automatic):
  1. /home/ubuntu/shared/allocator_history.json — daily equity per bot
  2. Alpaca fill history per bot account (keys from the allocator's .env)
     -> realized P/L per ticker per bot, FIFO-matched, options mapped to
        their underlying. No log parsing; format-proof.

Rebuilds /home/ubuntu/fleet_stats.xlsx from scratch every run:
  Dashboard  — KPI cards, leaderboard, equity curves, return bars
  Daily P/L  — per-bot per-day dollar change heat map + fleet daily totals
  Tickers    — per bot: every traded symbol ranked by realized P/L
  Data       — raw equity matrix

Cron: 0 1 * * * /usr/bin/python3 /home/ubuntu/fleetstats/fleet_stats.py >> /home/ubuntu/fleetstats/fleet_stats.log 2>&1
"""
import json
import re
from collections import defaultdict, deque
from datetime import datetime
from pathlib import Path

import requests
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HISTORY = Path("/home/ubuntu/shared/allocator_history.json")
ALLOC_ENV = Path("/home/ubuntu/allocator/.env")
OUT = Path("/home/ubuntu/fleet_stats.xlsx")
ALPACA = "https://paper-api.alpaca.markets"

ENV_PREFIX = {
    "Trend Confirmation": "TREND", "Mean Reversion": "MEANREV", "ORB Breakout": "ORB",
    "Overnight Drift": "OVERNIGHT", "Momentum Rotation": "MOMENTUM", "Credit Spreads": "CREDIT",
    "Defensive Rotation": "DEFENSE", "Pairs Rel Value": "PAIRS", "Passive Benchmark": "BENCHMARK",
    "Trend Regime": "TRENDREGIME", "VWAP MeanRev": "VWAP", "Gap Bot": "GAP", "VRP Bot": "VRP",
}
KILL = {
    "Trend Confirmation": (25.0, -20.0), "Mean Reversion": (15.0, -12.0),
    "ORB Breakout": (20.0, -15.0), "Overnight Drift": (15.0, -12.0),
    "Momentum Rotation": (20.0, -15.0), "Credit Spreads": (15.0, -12.0),
    "Defensive Rotation": (20.0, -15.0), "Pairs Rel Value": (15.0, -12.0),
    "Passive Benchmark": (40.0, -40.0), "Trend Regime": (25.0, -20.0),
    "VWAP MeanRev": (15.0, -12.0), "Gap Bot": (15.0, -12.0), "VRP Bot": (15.0, -12.0),
}

NAVY = "102A43"; NAVY2 = "1B3A5C"; WHITE = "F0F4F8"
GREEN = "2ECC71"; RED = "E74C3C"
F_TITLE = Font(size=22, bold=True, color=WHITE)
F_SUB = Font(size=11, color="BCCCDC")
F_KPI_L = Font(size=10, bold=True, color=WHITE)
F_KPI_V = Font(size=16, bold=True, color=WHITE)
F_HDR = Font(size=11, bold=True, color=WHITE)
F_CELL = Font(size=11)
THIN = Border(*[Side(style="thin", color="D9E2EC")] * 4)
CENTER = Alignment(horizontal="center", vertical="center")


def fill(h): return PatternFill("solid", fgColor=h)


def heat(v, scale):
    """Green->white->red fill scaled to +/-scale dollars."""
    if v is None: return fill("FFFFFF")
    x = max(-1.0, min(1.0, v / scale))
    if x >= 0:
        g = int(255 - 100 * x)
        return fill(f"{int(213-160*x):02X}{245:02X}{int(227-100*x):02X}".upper()) if False else fill(
            "%02X%02X%02X" % (int(255 - 120 * x), 255 - int(20 * x), int(255 - 120 * x)))
    x = -x
    return fill("%02X%02X%02X" % (255, int(255 - 130 * x), int(255 - 130 * x)))


def load_env():
    env = {}
    if ALLOC_ENV.exists():
        for line in ALLOC_ENV.read_text().splitlines():
            line = line.strip()
            if line and not line.startswith("#") and "=" in line:
                k, v = line.split("=", 1)
                env[k.strip()] = v.strip()
    return env


def occ_underlying(symbol):
    m = re.match(r"^([A-Z]+)\d{6}[CP]\d{8}$", symbol)
    return (m.group(1), 100) if m else (symbol, 1)


def fifo_realized(fills):
    """FIFO-match a chronological fill list [(side, qty, price, mult)] -> realized $."""
    lots = deque()  # (signed_qty, price, mult)
    realized = 0.0
    for side, qty, price, mult in fills:
        q = qty if side == "buy" else -qty
        while q and lots and (lots[0][0] > 0) != (q > 0):
            lot_q, lot_p, m = lots[0]
            matched = min(abs(q), abs(lot_q))
            if lot_q > 0:   # closing a long with a sell
                realized += (price - lot_p) * matched * m
            else:           # covering a short with a buy
                realized += (lot_p - price) * matched * m
            lot_q += matched if lot_q < 0 else -matched
            q += matched if q < 0 else -matched
            if lot_q == 0: lots.popleft()
            else: lots[0] = (lot_q, lot_p, m)
        if q: lots.append((q, price, mult))
    return realized


def fetch_ticker_pnl(env):
    """Per bot: {underlying: realized $} from Alpaca fill activities."""
    out = {}
    for bot, prefix in ENV_PREFIX.items():
        key, sec = env.get(f"{prefix}_KEY"), env.get(f"{prefix}_SECRET")
        if not key or not sec:
            continue
        headers = {"APCA-API-KEY-ID": key, "APCA-API-SECRET-KEY": sec}
        fills, token = [], None
        try:
            acct = requests.get(f"{ALPACA}/v2/account", headers=headers, timeout=30).json()
            acct_id = str(acct.get("account_number", "?"))[-6:]
            print(f"  [{bot}] account ...{acct_id}, equity {acct.get('equity')}")
            for _ in range(20):  # pagination guard
                params = {"activity_types": "FILL", "page_size": 100, "direction": "asc"}
                if token: params["page_token"] = token
                r = requests.get(f"{ALPACA}/v2/account/activities", headers=headers,
                                 params=params, timeout=30)
                r.raise_for_status()
                batch = r.json()
                fills.extend(batch)
                if len(batch) < 100: break
                token = batch[-1]["id"]
        except Exception as exc:
            print(f"  [{bot}] activities fetch failed: {exc}")
            continue
        per_symbol = defaultdict(list)
        for f in sorted(fills, key=lambda x: x.get("transaction_time", "")):
            und, mult = occ_underlying(f["symbol"])
            per_symbol[und].append((f["side"].replace("sell_short", "sell"),
                                    float(f["qty"]), float(f["price"]), mult))
        out[bot] = {s: fifo_realized(fl) for s, fl in per_symbol.items()}
        print(f"  [{bot}] {len(fills)} fills across {len(per_symbol)} symbols")
    return out


def stats_for(points):
    eq = [v for _, v in points]
    start, cur = eq[0], eq[-1]
    ret = (cur / start - 1) * 100
    peak, maxdd = eq[0], 0.0
    for v in eq:
        peak = max(peak, v)
        maxdd = min(maxdd, (v / peak - 1) * 100)
    return start, cur, ret, maxdd, len(eq)


def main():
    hist = json.loads(HISTORY.read_text())
    dates = sorted(hist.keys())
    bots = list(ENV_PREFIX.keys())
    bots = [b for b in bots if any(b in hist[d] for d in dates)] + \
           [b for d in dates for b in hist[d] if b not in ENV_PREFIX]
    series = {b: [(d, hist[d][b]) for d in dates if b in hist[d]] for b in bots}

    rows = []
    for b in bots:
        start, cur, ret, maxdd, days = stats_for(series[b])
        ddl, fll = KILL.get(b, (20.0, -15.0))
        headroom = min(ddl + maxdd, ret - fll)
        rows.append([b, cur, ret, maxdd, days, f"{ddl:.0f}% / {fll:.0f}%", headroom])
    rows.sort(key=lambda r: -r[2])
    fleet_cur = sum(r[1] for r in rows)
    fleet_start = sum(stats_for(series[b])[0] for b in bots)
    fleet_ret = (fleet_cur / fleet_start - 1) * 100
    best, worst = rows[0], rows[-1]

    print("Fetching per-ticker P/L from Alpaca fill history (13 accounts)...")
    tickers = fetch_ticker_pnl(load_env())

    wb = Workbook()

    # ================= DASHBOARD =================
    ws = wb.active; ws.title = "Dashboard"; ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    for c in "BCDEFGHIJKL": ws.column_dimensions[c].width = 15
    for r in range(1, 5):
        for c in range(1, 13): ws.cell(row=r, column=c).fill = fill(NAVY)
    ws.merge_cells("B2:H2"); ws["B2"] = "FLEET COMMAND"; ws["B2"].font = F_TITLE
    ws.merge_cells("B3:L3")
    ws["B3"] = f"13-bot paper fleet · rebuilt {datetime.now():%Y-%m-%d %H:%M} UTC · equity from allocator · tickers from Alpaca fills"
    ws["B3"].font = F_SUB

    kpis = [("FLEET VALUE", f"${fleet_cur:,.0f}", NAVY2),
            ("FLEET RETURN", f"{fleet_ret:+.2f}%", GREEN if fleet_ret >= 0 else RED),
            ("BEST BOT", f"{best[0]}", GREEN), ("", f"{best[2]:+.2f}%", GREEN),
            ("WORST BOT", f"{worst[0]}", RED if worst[2] < 0 else NAVY2),
            ("", f"{worst[2]:+.2f}%", RED if worst[2] < 0 else NAVY2)]
    # 3 wide cards: value, return, best, worst (name on row 7, pct row 8)
    cards = [("FLEET VALUE", f"${fleet_cur:,.0f}", None, NAVY2),
             ("FLEET RETURN", f"{fleet_ret:+.2f}%", None, GREEN if fleet_ret >= 0 else RED),
             ("BEST BOT", best[0], f"{best[2]:+.2f}%", GREEN),
             ("WORST BOT", worst[0], f"{worst[2]:+.2f}%", RED if worst[2] < 0 else NAVY2)]
    col = 2
    for label, line1, line2, color in cards:
        ws.merge_cells(start_row=6, start_column=col, end_row=6, end_column=col + 1)
        ws.merge_cells(start_row=7, start_column=col, end_row=7, end_column=col + 1)
        ws.merge_cells(start_row=8, start_column=col, end_row=8, end_column=col + 1)
        for r in (6, 7, 8):
            for c in (col, col + 1): ws.cell(row=r, column=c).fill = fill(color)
        ws.cell(row=6, column=col, value=label).font = F_KPI_L
        c1 = ws.cell(row=7, column=col, value=line1); c1.font = F_KPI_V; c1.alignment = CENTER
        if line2:
            c2 = ws.cell(row=8, column=col, value=line2)
            c2.font = Font(size=13, bold=True, color=WHITE); c2.alignment = CENTER
        col += 3
    ws.row_dimensions[7].height = 24

    ws["B10"] = "LEADERBOARD"; ws["B10"].font = Font(size=14, bold=True, color=NAVY)
    hdr = ["Bot", "Equity", "Return %", "Max DD %", "Days", "Kill-Lines", "Headroom %"]
    for i, h in enumerate(hdr):
        c = ws.cell(row=11, column=2 + i, value=h)
        c.font = F_HDR; c.fill = fill(NAVY); c.alignment = CENTER
    r = 12
    for name, cur, ret, maxdd, days, kl, headroom in rows:
        for i, v in enumerate([name, cur, ret, maxdd, days, kl, headroom]):
            c = ws.cell(row=r, column=2 + i, value=v); c.font = F_CELL; c.border = THIN
            if i == 1: c.number_format = "$#,##0"
            if i in (2, 3, 6): c.number_format = "+0.00;-0.00"
        rc = ws.cell(row=r, column=4)
        rc.fill = fill("D5F5E3" if ret >= 0 else "FADBD8")
        rc.font = Font(bold=True, color="1E8449" if ret >= 0 else "943126")
        hr = ws.cell(row=r, column=8)
        hr.fill = fill("FADBD8" if headroom < 5 else ("FDEBD0" if headroom < 10 else "D5F5E3"))
        r += 1

    # ================= DATA =================
    wd = wb.create_sheet("Data")
    wd["A1"] = "Date"
    for j, b in enumerate(bots): wd.cell(row=1, column=2 + j, value=b)
    for i, d in enumerate(dates):
        wd.cell(row=2 + i, column=1, value=d)
        for j, b in enumerate(bots):
            if b in hist[d]: wd.cell(row=2 + i, column=2 + j, value=hist[d][b])
    wd.column_dimensions["A"].width = 12
    for j in range(len(bots)): wd.column_dimensions[get_column_letter(2 + j)].width = 20

    # ================= CHARTS (readability tuned) =================
    all_eq = [v for b in bots for _, v in series[b]]
    lo, hi = min(all_eq), max(all_eq)
    pad = max((hi - lo) * 0.15, 200)
    lc = LineChart()
    lc.title = "Equity Curves"
    lc.height, lc.width, lc.style = 13, 32, 12
    lc.y_axis.title = "Equity ($)"
    lc.y_axis.scaling.min = lo - pad
    lc.y_axis.scaling.max = hi + pad
    lc.y_axis.numFmt = "$#,##0"
    lc.legend.position = "b"
    lc.add_data(Reference(wd, min_col=2, max_col=1 + len(bots), min_row=1, max_row=1 + len(dates)),
                titles_from_data=True)
    lc.set_categories(Reference(wd, min_col=1, min_row=2, max_row=1 + len(dates)))
    ws.add_chart(lc, f"B{r + 2}")

    # ================= DAILY P/L =================
    wp = wb.create_sheet("Daily P&L")
    wp.sheet_view.showGridLines = False
    wp.column_dimensions["A"].width = 20
    wp["A1"] = "DAILY P&L ($) — equity change per TRADING SESSION (allocator records at 8:16 PM ET)"
    wp["A1"].font = Font(size=14, bold=True, color=NAVY)
    from datetime import timedelta as _td
    def session_label(record_date):
        dt = datetime.strptime(record_date, "%Y-%m-%d") - _td(days=1)
        return f"Session {dt:%m-%d}"
    for j, d in enumerate(dates[1:] if len(dates) > 1 else dates):
        c = wp.cell(row=2, column=2 + j, value=session_label(d))
        c.font = F_HDR; c.fill = fill(NAVY); c.alignment = CENTER
        wp.column_dimensions[get_column_letter(2 + j)].width = 13
    tot_col = 2 + max(len(dates) - 1, 1)
    tc = wp.cell(row=2, column=tot_col, value="TOTAL")
    tc.font = F_HDR; tc.fill = fill(NAVY2); tc.alignment = CENTER
    wp.column_dimensions[get_column_letter(tot_col)].width = 13
    fleet_daily = defaultdict(float)
    for i, b in enumerate([row_[0] for row_ in rows]):
        wp.cell(row=3 + i, column=1, value=b).font = F_CELL
        eq_by_date = dict(series[b])
        total = 0.0
        for j in range(1, len(dates)):
            d0, d1 = dates[j - 1], dates[j]
            if d0 in eq_by_date and d1 in eq_by_date:
                delta = eq_by_date[d1] - eq_by_date[d0]
                fleet_daily[d1] += delta; total += delta
                c = wp.cell(row=3 + i, column=1 + j, value=round(delta, 2))
                c.number_format = "+$#,##0;-$#,##0"; c.fill = heat(delta, 800); c.border = THIN
        c = wp.cell(row=3 + i, column=tot_col, value=round(total, 2))
        c.number_format = "+$#,##0;-$#,##0"; c.font = Font(bold=True)
        c.fill = fill("D5F5E3" if total >= 0 else "FADBD8")
    fr = 3 + len(rows)
    wp.cell(row=fr, column=1, value="FLEET").font = Font(bold=True, color=NAVY)
    for j in range(1, len(dates)):
        d1 = dates[j]
        c = wp.cell(row=fr, column=1 + j, value=round(fleet_daily[d1], 2))
        c.number_format = "+$#,##0;-$#,##0"; c.font = Font(bold=True); c.fill = heat(fleet_daily[d1], 2500)
    c = wp.cell(row=fr, column=tot_col, value=round(sum(fleet_daily.values()), 2))
    c.number_format = "+$#,##0;-$#,##0"; c.font = Font(bold=True, color=NAVY)

    # ================= TICKERS =================
    wt = wb.create_sheet("Tickers")
    wt.sheet_view.showGridLines = False
    wt.column_dimensions["A"].width = 20; wt.column_dimensions["B"].width = 12
    wt.column_dimensions["C"].width = 14
    wt["A1"] = "REALIZED P/L BY TICKER — from Alpaca fill history (open positions not included)"
    wt["A1"].font = Font(size=14, bold=True, color=NAVY)
    r = 3
    if not tickers:
        wt["A3"] = "No fill data available (keys missing or API unreachable this run)."
    for bot in [row_[0] for row_ in rows]:
        if bot not in tickers: continue
        sym_pnl = tickers[bot]
        wt.cell(row=r, column=1, value=bot).font = Font(size=12, bold=True, color=WHITE)
        for c in range(1, 4): wt.cell(row=r, column=c).fill = fill(NAVY)
        r += 1
        shown = {s: p for s, p in sym_pnl.items() if abs(p) >= 0.01}
        if not shown:
            note = "(no closed trades yet)" if not sym_pnl else "(positions open, nothing realized yet)"
            wt.cell(row=r, column=1, value=note).font = Font(italic=True, color="8492A6")
            r += 2
            continue
        for sym, pnl in sorted(shown.items(), key=lambda t: -t[1]):
            wt.cell(row=r, column=1, value=sym).font = F_CELL
            c = wt.cell(row=r, column=2, value=round(pnl, 2))
            c.number_format = "+$#,##0;-$#,##0"
            c.fill = fill("D5F5E3" if pnl >= 0 else "FADBD8")
            c.font = Font(bold=True, color="1E8449" if pnl >= 0 else "943126")
            c.border = THIN
            r += 1
        r += 1

    wb.save(OUT)
    print(f"{datetime.now():%Y-%m-%d %H:%M:%S} rebuilt: {len(bots)} bots, {len(dates)} days, "
          f"fleet {fleet_ret:+.2f}%, ticker data for {len(tickers)} bots -> {OUT}")


if __name__ == "__main__":
    main()
